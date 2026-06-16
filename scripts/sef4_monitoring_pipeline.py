import os
import requests
import pandas as pd

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
# =====================================
# CONFIG
# =====================================

TENANT_ID = os.getenv("AZURE_TENANT_ID")
CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

SHAREPOINT_SITE = "thelearningtrust.sharepoint.com"
SITE_PATH = "/sites/TheLearningTrust"

WORKBOOK_NAME = "SEF 4 Monthly Stats Analysis.xlsx"

MONTHLY_REPORT_ID = "1itWYfUTkxA3g1qJThcaQ9XmgkOzOrngMgqjNSess20A"
MONTHLY_SHEET = "Monthly reporting"


# =====================================
# MICROSOFT GRAPH AUTH
# =====================================

def get_access_token():

    url = (
        f"https://login.microsoftonline.com/"
        f"{TENANT_ID}/oauth2/v2.0/token"
    )

    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }

    r = requests.post(url, data=data)
    r.raise_for_status()

    return r.json()["access_token"]


def get_drive_id():

    token = get_access_token()

    headers = {
        "Authorization": f"Bearer {token}"
    }

    site_url = (
        f"https://graph.microsoft.com/v1.0/sites/"
        f"{SHAREPOINT_SITE}:{SITE_PATH}"
    )

    site_res = requests.get(site_url, headers=headers)
    site_res.raise_for_status()

    site_id = site_res.json()["id"]

    drive_url = (
        f"https://graph.microsoft.com/v1.0/sites/"
        f"{site_id}/drives"
    )

    drive_res = requests.get(drive_url, headers=headers)
    drive_res.raise_for_status()

    drives = drive_res.json()["value"]

    for drive in drives:

        if drive["name"] in ["Documents", "Shared Documents"]:
            return drive["id"]

    raise Exception("Documents library not found")


# =====================================
# DOWNLOAD WORKBOOK
# =====================================

def download_from_sharepoint(filename):

    token = get_access_token()
    drive_id = get_drive_id()

    headers = {
        "Authorization": f"Bearer {token}"
    }

    search_url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        f"/root:/Consolidated data:/children"
    )

    res = requests.get(search_url, headers=headers)
    res.raise_for_status()

    file_id = None

    for item in res.json()["value"]:
        print("FOUND:", item["name"])

        if item["name"].strip() == filename.strip():
            file_id = item["id"]
            break

    if not file_id:
        raise Exception(f"{filename} not found")

    download_url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        f"/items/{file_id}/content"
    )

    r = requests.get(download_url, headers=headers)
    r.raise_for_status()

    with open(filename, "wb") as f:
        f.write(r.content)

    print(f"✅ Downloaded {filename}")
# =====================================
# UPLOAD WORKBOOK
# =====================================

def upload_to_sharepoint(filename):

    token = get_access_token()
    drive_id = get_drive_id()

    upload_url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        f"/root:/Consolidated data/{filename}:/content"
    )

    with open(filename, "rb") as f:

        res = requests.put(
            upload_url,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type":
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
            data=f
        )

    res.raise_for_status()

    print("✅ Uploaded workbook")
# =====================================
# LOAD MAPPINGS
# =====================================

def load_mappings(workbook_name):

    mapping_df = pd.read_excel(
        workbook_name,
        sheet_name="variables to be changed",
        header=None
    )

    print("MAPPING SHAPE:", mapping_df.shape)

    ip_mapping = {}

    for _, row in mapping_df.iloc[3:50].iterrows():

        try:

            source = row.iloc[2]
            target = row.iloc[3]

            if pd.notna(source) and pd.notna(target):
                ip_mapping[
                    str(source).strip()
                ] = str(target).strip()

        except Exception:
            pass

    indicator_mapping = {}

    for _, row in mapping_df.iloc[3:50].iterrows():

        try:

            source = row.iloc[7]
            target = row.iloc[8]

            if pd.notna(source) and pd.notna(target):

                source = (
                    str(source)
                    .replace("\n", " ")
                    .strip()
                )

                target = (
                    str(target)
                    .replace("\n", " ")
                    .strip()
                )

                indicator_mapping[source] = target

        except Exception:
            pass

    print("IP MAPPINGS:", len(ip_mapping))
    print("INDICATOR MAPPINGS:", len(indicator_mapping))

    return ip_mapping, indicator_mapping
# =====================================
# MONTHLY REPORT
# =====================================

def build_outputs(ip_mapping, indicator_mapping):

    monthly_url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{MONTHLY_REPORT_ID}/export?format=xlsx"
    )

    raw = pd.read_excel(
        monthly_url,
        sheet_name="Monthly reporting",
        header=None
    )

    # Row 2 contains actual indicator names
    headers = (
        raw.iloc[1]
        .fillna("")
        .astype(str)
        .str.replace("\n", " ", regex=False)
        .str.strip()
    )

    df = raw.iloc[2:].copy()
    df.columns = headers

    # Remove empty rows
    df = df.dropna(how="all")

    # Column B = ADC / Organisation
    ip_col = df.columns[1]

    print(f"\nIP Column Found: {ip_col}")

    # Clean column names
    df.columns = [
        str(c).replace("\n", " ").strip()
        for c in df.columns
    ]

    # Clean mapping keys
    valid_indicators = {
        str(k).replace("\n", " ").strip()
        for k in indicator_mapping.keys()
    }

    # ONLY KEEP THE 21 INDICATORS FROM MAPPING FILE
    indicator_cols = []

    for col in df.columns:

        clean_col = str(col).strip()

        if clean_col in valid_indicators:
            indicator_cols.append(col)

    print("\n===== INDICATORS USED =====")

    for col in indicator_cols:
        print(col)

    print(
        f"\nTotal indicators found: "
        f"{len(indicator_cols)}"
    )

    print("\n===== INDICATORS NOT FOUND =====")

    for indicator in valid_indicators:

        if indicator not in [
            str(c).strip()
            for c in indicator_cols
        ]:

            print(f"❌ {indicator}")

    # Convert indicators to numeric
    for col in indicator_cols:

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        ).fillna(0)

    # Aggregate by organisation
    outputs = (
        df
        .groupby(ip_col, dropna=True)[indicator_cols]
        .sum()
        .reset_index()
    )

    outputs.rename(
        columns={ip_col: "IP Name"},
        inplace=True
    )

    outputs["IP Name"] = (
        outputs["IP Name"]
        .astype(str)
        .str.strip()
    )

    # Apply organisation mapping
    outputs["IP Name"] = (
        outputs["IP Name"]
        .replace(ip_mapping)
    )

    # Rename indicators to expected output names
    outputs.rename(
        columns=indicator_mapping,
        inplace=True
    )

    outputs["Criteria"] = "Outputs"

    print("\n===== OUTPUT ORGANISATIONS =====")
    print(outputs["IP Name"].tolist())

    print("\n===== OUTPUT SHAPE =====")
    print(outputs.shape)

    return outputs
# =====================================
# ADDITIONAL TABS
# =====================================
def build_monitoring_outputs(
    ip_mapping,
    indicator_mapping
):

    monthly_url = (
        f"https://docs.google.com/spreadsheets/d/"
        f"{MONTHLY_REPORT_ID}/export?format=xlsx"
    )

    raw = pd.read_excel(
        monthly_url,
        sheet_name="Monthly reporting",
        header=None
    )

    headers = (
        raw.iloc[1]
        .fillna("")
        .astype(str)
        .str.replace("\n", " ")
        .str.strip()
    )

    df = raw.iloc[2:].copy()

    df.columns = headers

    df = df.dropna(
        how="all"
    )

    month_col = df.columns[0]
    ip_col = df.columns[1]

    valid_indicators = {
        str(k).replace("\n", " ").strip()
        for k in indicator_mapping.keys()
    }

    indicator_cols = []

    for col in df.columns:

        if str(col).strip() in valid_indicators:
            indicator_cols.append(col)

    for col in indicator_cols:

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        ).fillna(0)

    df[ip_col] = (
        df[ip_col]
        .astype(str)
        .str.strip()
        .replace(ip_mapping)
    )

    # ==========================
    # OUTPUT 1
    # CUMULATIVE
    # ==========================

    outputs_cumulative = (
        df
        .groupby(ip_col)[indicator_cols]
        .sum()
        .reset_index()
    )

    outputs_cumulative.rename(
        columns={ip_col: "IP Name"},
        inplace=True
    )

    outputs_cumulative.rename(
        columns=indicator_mapping,
        inplace=True
    )

    outputs_cumulative["Criteria"] = "Outputs"

    # ==========================
    # OUTPUT 2
    # MONTHLY
    # ==========================

    outputs_monthly = (
        df
        .groupby(
            [ip_col, month_col]
        )[indicator_cols]
        .sum()
        .reset_index()
    )

    outputs_monthly.rename(
        columns={
            ip_col: "IP Name",
            month_col: "Month"
        },
        inplace=True
    )

    outputs_monthly.rename(
        columns=indicator_mapping,
        inplace=True
    )

    # ==========================
    # OUTPUT 3
    # RAW HISTORICAL
    # ==========================

    raw_historical = outputs_monthly.melt(
        id_vars=[
            "IP Name",
            "Month"
        ],
        var_name="Indicator",
        value_name="Value"
    )

    # ==========================
    # OUTPUT 4
    # VARIANCE
    # ==========================

    month_order = [
        "November",
        "December",
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July"
    ]

    current_month = datetime.now().strftime("%B")

    if current_month in month_order:

        valid_months = month_order[
            :month_order.index(current_month) + 1
        ]

        outputs_monthly = outputs_monthly[
            outputs_monthly["Month"]
            .isin(valid_months)
        ]

    outputs_monthly["Month"] = pd.Categorical(
        outputs_monthly["Month"],
        categories=month_order,
        ordered=True
    )

    outputs_monthly = outputs_monthly.sort_values(
        ["IP Name", "Month"]
    )

    variance_rows = []

    indicators = [
        c
        for c in outputs_monthly.columns
        if c not in [
            "IP Name",
            "Month"
        ]
    ]

    for ip in outputs_monthly["IP Name"].unique():

        ip_data = outputs_monthly[
            outputs_monthly["IP Name"] == ip
        ]

        for indicator in indicators:

            previous = None

            for _, row in ip_data.iterrows():

                current = row[indicator]

                if previous is not None:

                    change = current - previous

                    if previous == 0:
                        pct_change = None
                    else:
                        pct_change = (
                            change / previous
                        ) * 100

                    status = "Normal"

                    if pct_change is not None:

                        if abs(pct_change) >= 100:
                            status = "Critical"

                        elif abs(pct_change) >= 50:
                            status = "Review"

                        elif abs(pct_change) >= 25:
                            status = "Monitor"

                    variance_rows.append(
                        {
                            "IP Name": ip,
                            "Month": row["Month"],
                            "Indicator": indicator,
                            "Previous": previous,
                            "Current": current,
                            "Change": change,
                            "Percent Change": pct_change,
                            "Status": status
                        }
                    )

                previous = current

    outputs_variance = pd.DataFrame(
        variance_rows
    )
    
# ==========================
# OUTPUT 5
# QUALITY CHECKS
# ==========================


quality_checks = outputs_variance[
outputs_variance["Status"] != "Normal"
].copy()

variance_heatmap = build_variance_heatmap(
    outputs_variance
    )

ip_heatmap = build_ip_heatmap(
        outputs_variance
    )

 top_increases, top_decreases = (
        build_executive_summary(
            outputs_variance
        )
    )

    return (
        outputs_cumulative,
        outputs_monthly,
        outputs_variance,
        quality_checks,
        raw_historical,
        variance_heatmap,
        ip_heatmap,
        top_increases,
        top_decreases
    )

# ==========================
# OUTPUT 6
# heatmap dataset
# ==========================
def build_variance_heatmap(outputs_variance):

    heatmap = (
        outputs_variance
        .pivot_table(
            index=["IP Name","Indicator"],
            columns="Month",
            values="Percent Change",
            aggfunc="sum"
        )
        .fillna(0)
        .reset_index()
    )

    month_cols = [
        c for c in heatmap.columns
        if c not in ["IP Name","Indicator"]
    ]

    heatmap["Grand Total"] = (
        heatmap[month_cols]
        .sum(axis=1)
    )

    return heatmap
def build_ip_heatmap(outputs_variance):

    heatmap = (
        outputs_variance
        .pivot_table(
            index="IP Name",
            columns="Month",
            values="Percent Change",
            aggfunc="sum"
        )
        .fillna(0)
        .reset_index()
    )

    month_cols = [
        c for c in heatmap.columns
        if c != "IP Name"
    ]

    heatmap["Grand Total"] = (
        heatmap[month_cols]
        .sum(axis=1)
    )

    return heatmap
# =====================================
# executive summary
# =====================================
def build_executive_summary(outputs_variance):

    summary = (
        outputs_variance
        .sort_values(
            "Percent Change",
            ascending=False
        )
        .copy()
    )

    top_increases = summary.head(20)

    top_decreases = summary.tail(20)

    return top_increases, top_decreases

# =====================================
# UPDATE WORKBOOK
# =====================================
def update_monitoring_workbook(
    workbook_name,
    outputs_cumulative,
    outputs_monthly,
    outputs_variance,
    quality_checks,
    raw_historical,
    variance_heatmap,
    ip_heatmap,
    top_increases,
    top_decreases
):

    with pd.ExcelWriter(
        workbook_name,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:

        outputs_cumulative.to_excel(
            writer,
            sheet_name="DATASOURCE",
            index=False
        )

        outputs_monthly.to_excel(
            writer,
            sheet_name="Outputs_Monthly",
            index=False
        )

        outputs_variance.to_excel(
            writer,
            sheet_name="Outputs_Variance",
            index=False
        )

        quality_checks.to_excel(
            writer,
            sheet_name="Quality_Checks",
            index=False
        )

        raw_historical.to_excel(
            writer,
            sheet_name="Raw_Historical_Data",
            index=False
        )
                variance_heatmap.to_excel(
            writer,
            sheet_name="Variance_Heatmap",
            index=False
        )

        ip_heatmap.to_excel(
            writer,
            sheet_name="IP_Heatmap",
            index=False
        )

        top_increases.to_excel(
            writer,
            sheet_name="Top_Increases",
            index=False
        )

        top_decreases.to_excel(
            writer,
            sheet_name="Top_Decreases",
            index=False
        )
    wb = load_workbook(workbook_name)

    ws = wb["Variance_Heatmap"]

    red_fill = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid"
    )

    ws.conditional_formatting.add(
        "C2:Z5000",
        CellIsRule(
            operator="greaterThan",
            formula=["100"],
            fill=red_fill
        )
    )

    ws.conditional_formatting.add(
        "C2:Z5000",
        CellIsRule(
            operator="lessThan",
            formula=["-100"],
            fill=red_fill
        )
    )

    wb.save(workbook_name)
    print(
        "Monitoring workbook updated"
    )
# =====================================
# MAIN
# =====================================

def run_pipeline():

    print("Starting SEF4 Monitoring Pipeline")

    download_from_sharepoint(
        WORKBOOK_NAME
    )

    ip_mapping, indicator_mapping = load_mappings(
        WORKBOOK_NAME
    )

   (
    outputs_cumulative,
    outputs_monthly,
    outputs_variance,
    quality_checks,
    raw_historical,
    variance_heatmap,
    ip_heatmap,
    top_increases,
    top_decreases
    ) = build_monitoring_outputs(
        ip_mapping,
        indicator_mapping
    )

    update_monitoring_workbook(
    WORKBOOK_NAME,
    outputs_cumulative,
    outputs_monthly,
    outputs_variance,
    quality_checks,
    raw_historical,
    variance_heatmap,
    ip_heatmap,
    top_increases,
    top_decreases
)

    upload_to_sharepoint(
        WORKBOOK_NAME
    )

    print(
        "SEF4 Monitoring Pipeline Complete"
    )
if __name__ == "__main__":

    print("SCRIPT EXECUTION STARTED")

    run_pipeline()
